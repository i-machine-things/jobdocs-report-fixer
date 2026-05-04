"""
Report Fixer Module - Transform Excel Reports to Match Template Layout

This module transforms Excel job reports (e.g., retech_jobRpt.xls) to match
a template layout. It tracks schedule changes, adds notes for date modifications,
and exports formatted Excel files with highlighting.
"""

import re
import sys
import json
from pathlib import Path
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from PyQt6.QtWidgets import (
    QWidget, QTableWidgetItem, QFileDialog, QHeaderView, QAbstractItemView,
    QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6 import uic

from core.base_module import BaseModule
from shared.utils import get_config_dir, open_folder

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportingModule(BaseModule):
    """Module for fixing and transforming Excel job reports"""

    def __init__(self):
        super().__init__()
        self._widget = None

        # Widget references
        self.template_path_edit = None
        self.source_path_edit = None
        self.source_info_label = None
        self.delivery_path_edit = None
        self.delivery_info_label = None
        self.customer_combo = None
        self.preview_table = None
        self.mapping_status_label = None
        self.fix_export_btn = None
        self.open_output_btn = None
        self.preview_customers_btn = None
        self.status_text = None

        # State
        self.source_df = None
        self.delivery_df = None
        self.template_columns = None
        self.last_output_path = None
        self.customer_column = None  # Detected customer column name
        self.available_customers = []  # List of customer folder names
        self.customer_mapping = {}  # Cached customer mapping from preview
        self.unmatched_customers = []  # Cached unmatched customers from preview
        self.preview_mode = 'columns'  # 'columns' or 'customers'
        self._aliases_cache = None  # Cached customer aliases (invalidated on save)

    def get_name(self) -> str:
        return "Report Fixer"

    def get_order(self) -> int:
        return 80

    def is_experimental(self) -> bool:
        return False

    def initialize(self, app_context):
        super().initialize(app_context)

    def get_widget(self) -> QWidget:
        if self._widget is None:
            self._widget = self._create_widget()
        return self._widget

    def _create_widget(self) -> QWidget:
        """Create the report fixer tab widget"""
        widget = QWidget()

        # Check dependencies
        if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            from PyQt6.QtWidgets import QVBoxLayout, QLabel
            layout = QVBoxLayout(widget)
            missing = []
            if not PANDAS_AVAILABLE:
                missing.append("pandas")
            if not OPENPYXL_AVAILABLE:
                missing.append("openpyxl")
            label = QLabel(f"Missing required packages: {', '.join(missing)}\n\n"
                          f"Install with: pip install {' '.join(missing)}")
            label.setStyleSheet("color: red; padding: 20px;")
            layout.addWidget(label)
            return widget

        # Load UI file
        ui_file = self._get_ui_path()
        uic.loadUi(ui_file, widget)

        # Store widget references
        self.template_path_edit = widget.template_path_edit
        self.source_path_edit = widget.source_path_edit
        self.source_info_label = widget.source_info_label
        self.delivery_path_edit = widget.delivery_path_edit
        self.delivery_info_label = widget.delivery_info_label
        self.customer_combo = widget.customer_combo
        self.preview_table = widget.preview_table
        self.mapping_status_label = widget.mapping_status_label
        self.fix_export_btn = widget.fix_export_btn
        self.open_output_btn = widget.open_output_btn
        self.preview_customers_btn = widget.preview_customers_btn
        self.status_text = widget.status_text

        # Setup table properties
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.preview_table.doubleClicked.connect(self._on_preview_double_click)

        # Connect signals
        widget.browse_template_btn.clicked.connect(self.browse_template)
        widget.browse_source_btn.clicked.connect(self.browse_source)
        widget.browse_delivery_btn.clicked.connect(self.browse_delivery)
        widget.fix_export_btn.clicked.connect(self.fix_and_export)
        widget.open_output_btn.clicked.connect(self.open_output_folder)
        widget.preview_customers_btn.clicked.connect(self.preview_customers)
        widget.customer_combo.currentTextChanged.connect(self._on_customer_changed)

        # Enable drag and drop on the widget
        widget.setAcceptDrops(True)
        widget.dragEnterEvent = self._drag_enter_event
        widget.dropEvent = self._drop_event

        # Load saved template path
        saved_template = self.app_context.get_setting('report_template_path', '')
        if saved_template and Path(saved_template).exists():
            self.template_path_edit.setText(saved_template)
            self._load_template(saved_template)

        # Load saved source path
        saved_source = self.app_context.get_setting('report_source_path', '')
        if saved_source and Path(saved_source).exists():
            self._load_source(saved_source)

        # Load saved delivery schedule path
        saved_delivery = self.app_context.get_setting('report_delivery_path', '')
        if saved_delivery and Path(saved_delivery).exists():
            self._load_delivery_schedule(saved_delivery)

        # Populate customer list
        self._populate_customers()

        return widget

    def _get_ui_path(self) -> Path:
        """Get path to the UI file (always relative to this plugin module)."""
        ui_file = Path(__file__).parent / 'ui' / 'reporting_tab.ui'
        if not ui_file.exists():
            raise FileNotFoundError(f"UI file not found: {ui_file}")
        return ui_file

    def _populate_customers(self):
        """Populate customer dropdown from customer_files_dir"""
        self.customer_combo.clear()
        self.available_customers = self.app_context.get_customer_list()
        self.customer_combo.addItem("-- Auto-detect from data --")
        self.customer_combo.addItems(sorted(self.available_customers))

    def _drag_enter_event(self, event):
        """Handle drag enter for file drops"""
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith(('.xls', '.xlsx')):
                    event.acceptProposedAction()
                    return
        event.ignore()

    def _drop_event(self, event):
        """Handle file drop"""
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(('.xls', '.xlsx')):
                self._load_source(file_path)
                self.app_context.set_setting('report_source_path', file_path)
                self.app_context.save_settings()
                break

    # ==================== File Loading ====================

    def browse_template(self):
        """Browse for template Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._widget,
            "Select Template File",
            "",
            "Excel Files (*.xls *.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.template_path_edit.setText(file_path)
            self._load_template(file_path)
            # Save to settings
            self.app_context.set_setting('report_template_path', file_path)
            self.app_context.save_settings()

    def browse_source(self):
        """Browse for source Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._widget,
            "Select Source Report File",
            "",
            "Excel Files (*.xls *.xlsx);;All Files (*.*)"
        )
        if file_path:
            self._load_source(file_path)
            self.app_context.set_setting('report_source_path', file_path)
            self.app_context.save_settings()

    def _load_template(self, file_path: str):
        """Load template file and extract column names"""
        try:
            df_template = pd.read_excel(file_path, nrows=0)
            self.template_columns = list(df_template.columns)
            self._log(f"Template loaded: {len(self.template_columns)} columns")
            self._update_preview()
        except Exception as e:
            self.show_error("Template Error", f"Failed to load template:\n{str(e)}")
            self.template_columns = None

    def _load_source(self, file_path: str):
        """Load source Excel file"""
        try:
            self.source_df = pd.read_excel(file_path)
            self.source_path_edit.setText(file_path)
            self.source_info_label.setText(
                f"Loaded {len(self.source_df)} rows x {len(self.source_df.columns)} columns"
            )
            self._log(f"Source loaded: {Path(file_path).name}")
            self._update_preview()
        except Exception as e:
            self.show_error("Source Error", f"Failed to load source file:\n{str(e)}")
            self.source_df = None
            self.source_info_label.setText("Failed to load file")

    def browse_delivery(self):
        """Browse for delivery schedule Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._widget,
            "Select Delivery Schedule File",
            "",
            "Excel Files (*.xls *.xlsx);;All Files (*.*)"
        )
        if file_path:
            self._load_delivery_schedule(file_path)
            self.app_context.set_setting('report_delivery_path', file_path)
            self.app_context.save_settings()

    def _load_delivery_schedule(self, file_path: str):
        """Load delivery schedule file — column A = Job ID, column F = Promise Date"""
        try:
            df = pd.read_excel(file_path, header=0)
            # Column F is index 5; rename to known names for merging
            cols = list(df.columns)
            if len(cols) < 6:
                self.show_error(
                    "Delivery Schedule Error",
                    f"File must have at least 6 columns (A–F). Found {len(cols)}."
                )
                self.delivery_df = None
                self.delivery_info_label.setText("Failed to load file")
                return
            job_col = cols[0]   # Column A: job number
            promise_col = cols[5]  # Column F: promise date

            self.delivery_df = df[[job_col, promise_col]].copy()
            self.delivery_df.columns = ['_delivery_job_id', 'Promise Date']
            # Normalize job ID to string for joining
            self.delivery_df['_delivery_job_id'] = (
                self.delivery_df['_delivery_job_id'].astype(str).str.strip()
            )
            # Convert promise date to date only
            self.delivery_df['Promise Date'] = pd.to_datetime(
                self.delivery_df['Promise Date'], errors='coerce'
            ).dt.date

            self.delivery_path_edit.setText(file_path)
            self.delivery_info_label.setText(
                f"Loaded {len(self.delivery_df)} rows — Promise Date from column F"
            )
            self._log(f"Delivery schedule loaded: {Path(file_path).name} ({len(self.delivery_df)} rows)")
        except Exception as e:
            self.show_error("Delivery Schedule Error", f"Failed to load delivery schedule:\n{str(e)}")
            self.delivery_df = None
            self.delivery_info_label.setText("Failed to load file")

    def _update_preview(self):
        """Update the column mapping preview table"""
        self.preview_mode = 'columns'
        self.preview_table.setRowCount(0)

        # Reset to 3 columns for column mapping view
        self.preview_table.setColumnCount(3)
        self.preview_table.setHorizontalHeaderLabels(["Status", "Column Name", "Action"])

        if self.template_columns is None or self.source_df is None:
            self.mapping_status_label.setText("Load template and source files to see column mapping")
            self.fix_export_btn.setEnabled(False)
            return

        source_cols = set(self.source_df.columns)
        template_cols = set(self.template_columns)

        matching = source_cols & template_cols
        to_remove = source_cols - template_cols
        to_add = template_cols - source_cols

        # Show all template columns with their status
        for col in self.template_columns:
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            if col in matching:
                status = "✓"
                action = "Keep (from source)"
                color = Qt.GlobalColor.darkGreen
            else:
                status = "+"
                action = "Add (empty column)"
                color = Qt.GlobalColor.blue

            status_item = QTableWidgetItem(status)
            status_item.setForeground(color)
            self.preview_table.setItem(row, 0, status_item)

            name_item = QTableWidgetItem(col)
            self.preview_table.setItem(row, 1, name_item)

            action_item = QTableWidgetItem(action)
            action_item.setForeground(color)
            self.preview_table.setItem(row, 2, action_item)

        # Show columns to be removed
        for col in sorted(to_remove):
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            status_item = QTableWidgetItem("✗")
            status_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 0, status_item)

            name_item = QTableWidgetItem(col)
            self.preview_table.setItem(row, 1, name_item)

            action_item = QTableWidgetItem("Remove (not in template)")
            action_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 2, action_item)

        self.mapping_status_label.setText(
            f"Matching: {len(matching)} | To Remove: {len(to_remove)} | To Add: {len(to_add)}"
        )

        # Enable buttons based on state
        self.fix_export_btn.setEnabled(True)
        self._update_button_states()

    def _on_customer_changed(self, text: str):
        """Handle customer dropdown selection change"""
        self._update_button_states()
        # Reset to column preview when customer changes
        if self.preview_mode == 'customers':
            self.preview_mode = 'columns'
            self._update_preview()

    def _update_button_states(self):
        """Update button enabled states based on current selection"""
        selected = self.customer_combo.currentText().strip()
        is_auto_detect = selected == "-- Auto-detect from data --"
        has_source = self.source_df is not None

        # Preview Customers button only enabled in auto-detect mode with source loaded
        self.preview_customers_btn.setEnabled(is_auto_detect and has_source)

        # Fix & Export enabled if we have data
        self.fix_export_btn.setEnabled(has_source and self.template_columns is not None)

    def preview_customers(self):
        """Preview customer matching before export"""
        if self.source_df is None:
            self.show_error("No Data", "Please load a source file first")
            return

        # Detect customer column
        customer_col = self._detect_customer_column()
        if not customer_col:
            self.show_error("No Customer Column",
                          "Could not detect a customer column in the source data.\n\n"
                          "Expected column names: Customer, Customer Name, Client, Company, etc.")
            return

        self._log("=" * 50)
        self._log(f"Previewing customer matching (column: '{customer_col}')")

        # Get unique customers from data
        unique_customers = self.source_df[customer_col].dropna().unique()

        # Match each customer to a folder
        self.customer_mapping = {}
        self.unmatched_customers = []

        for src_customer in unique_customers:
            src_name = str(src_customer).strip()
            if not src_name:
                continue

            matched_folder, score = self._fuzzy_match_customer(src_name)
            if matched_folder:
                # Count rows for this customer
                row_count = len(self.source_df[
                    self.source_df[customer_col].astype(str).str.strip() == src_name
                ])
                self.customer_mapping[src_name] = (matched_folder, score, row_count)
            else:
                row_count = len(self.source_df[
                    self.source_df[customer_col].astype(str).str.strip() == src_name
                ])
                self.unmatched_customers.append((src_name, row_count))

        # Update preview table to show customer matching
        self._show_customer_preview()

    def _show_customer_preview(self):
        """Display customer matching preview in the table"""
        self.preview_mode = 'customers'
        self.preview_table.setRowCount(0)

        # Update column headers for customer view
        self.preview_table.setColumnCount(4)
        self.preview_table.setHorizontalHeaderLabels(
            ["Status", "Source Customer", "Matched Folder", "Rows"]
        )

        # Show matched customers
        for src_name, (folder_name, score, row_count) in self.customer_mapping.items():
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            # Status based on match quality
            if score >= 0.95:
                status = "✓"
                color = Qt.GlobalColor.darkGreen
            elif score >= 0.8:
                status = "~"
                color = Qt.GlobalColor.darkYellow
            else:
                status = "?"
                color = Qt.GlobalColor.darkCyan

            status_item = QTableWidgetItem(status)
            status_item.setForeground(color)
            status_item.setToolTip(f"Match score: {score:.0%}")
            self.preview_table.setItem(row, 0, status_item)

            src_item = QTableWidgetItem(src_name)
            self.preview_table.setItem(row, 1, src_item)

            folder_item = QTableWidgetItem(folder_name)
            folder_item.setForeground(color)
            self.preview_table.setItem(row, 2, folder_item)

            rows_item = QTableWidgetItem(str(row_count))
            rows_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.preview_table.setItem(row, 3, rows_item)

            self._log(f"  ✓ '{src_name}' -> '{folder_name}' ({score:.0%}, {row_count} rows)")

        # Show unmatched customers
        for src_name, row_count in self.unmatched_customers:
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            status_item = QTableWidgetItem("✗")
            status_item.setForeground(Qt.GlobalColor.red)
            status_item.setToolTip("No matching folder found")
            self.preview_table.setItem(row, 0, status_item)

            src_item = QTableWidgetItem(src_name)
            src_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 1, src_item)

            folder_item = QTableWidgetItem("NO MATCH")
            folder_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 2, folder_item)

            rows_item = QTableWidgetItem(str(row_count))
            rows_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            rows_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 3, rows_item)

            self._log(f"  ✗ '{src_name}' -> NO MATCH ({row_count} rows will be skipped)")

        # Update status label
        matched_count = len(self.customer_mapping)
        unmatched_count = len(self.unmatched_customers)
        total_matched_rows = sum(info[2] for info in self.customer_mapping.values())
        total_unmatched_rows = sum(info[1] for info in self.unmatched_customers)

        self.mapping_status_label.setText(
            f"Matched: {matched_count} ({total_matched_rows} rows) | "
            f"Unmatched: {unmatched_count} ({total_unmatched_rows} rows)"
        )

        self._log("=" * 50)
        self._log(f"Preview complete: {matched_count} matched, {unmatched_count} unmatched")
        if unmatched_count > 0:
            self._log("TIP: Double-click unmatched rows to manually assign a folder")
        self._log("Click 'Fix & Export Report' to proceed with export")

    def _on_preview_double_click(self, index):
        """Handle double-click on preview table row to manually assign folder"""
        if self.preview_mode != 'customers':
            return

        row = index.row()
        # Get the source customer name from column 1
        src_item = self.preview_table.item(row, 1)
        if not src_item:
            return

        src_name = src_item.text()

        # Check if this is an unmatched customer
        folder_item = self.preview_table.item(row, 2)
        if not folder_item:
            return

        current_folder = folder_item.text()

        # Show dialog to select folder
        from PyQt6.QtWidgets import QInputDialog

        folders = sorted(self.available_customers)
        current_idx = folders.index(current_folder) if current_folder in folders else 0

        folder, ok = QInputDialog.getItem(
            self._widget,
            "Assign Customer Folder",
            f"Select folder for '{src_name}':",
            folders,
            current_idx,
            False  # Not editable
        )

        if ok and folder:
            # Update the mapping
            # Find row count for this customer
            row_count = 0
            if src_name in [name for name, _ in self.unmatched_customers]:
                # Was unmatched - find and remove from unmatched
                for i, (name, count) in enumerate(self.unmatched_customers):
                    if name == src_name:
                        row_count = count
                        self.unmatched_customers.pop(i)
                        break
                # Add to matched
                self.customer_mapping[src_name] = (folder, 1.0, row_count)
            else:
                # Was already matched - update
                if src_name in self.customer_mapping:
                    _, _, row_count = self.customer_mapping[src_name]
                    self.customer_mapping[src_name] = (folder, 1.0, row_count)

            # Save as alias for future use
            self._save_customer_alias(src_name, folder)

            # Refresh the preview
            self._show_customer_preview()

    # ==================== Fuzzy Matching ====================

    def _load_customer_aliases(self) -> dict:
        """Load manual customer name mappings from config file"""
        if self._aliases_cache is not None:
            return self._aliases_cache

        alias_file = get_config_dir() / 'customer_aliases.json'
        if alias_file.exists():
            try:
                with open(alias_file, 'r', encoding='utf-8') as f:
                    self._aliases_cache = json.load(f)
                    return self._aliases_cache
            except Exception as e:
                self._log(f"Warning: could not load customer aliases: {e}")
        self._aliases_cache = {}
        return self._aliases_cache

    def _save_customer_alias(self, source_name: str, folder_name: str):
        """Save a customer alias mapping"""
        alias_file = get_config_dir() / 'customer_aliases.json'
        aliases = self._load_customer_aliases()
        aliases[source_name.upper()] = folder_name
        try:
            with open(alias_file, 'w', encoding='utf-8') as f:
                json.dump(aliases, f, indent=2)
            self._aliases_cache = None  # Invalidate cache after write
            self._log(f"Saved alias: '{source_name}' -> '{folder_name}'")
        except Exception as e:
            self._log(f"Warning: Could not save alias: {e}")

    def _normalize_name(self, name: str) -> str:
        """Normalize customer name for comparison"""
        if not name:
            return ""
        # Lowercase, remove common suffixes, strip whitespace
        normalized = str(name).lower().strip()
        # Remove common business suffixes
        for suffix in [' inc', ' inc.', ' llc', ' llc.', ' corp', ' corp.',
                       ' co', ' co.', ' ltd', ' ltd.', ' company']:
            if normalized.endswith(suffix):
                normalized = normalized[:-len(suffix)].strip()
        # Remove punctuation and extra spaces
        normalized = ''.join(c if c.isalnum() or c.isspace() else ' ' for c in normalized)
        normalized = ' '.join(normalized.split())
        return normalized

    def _fuzzy_match_customer(self, source_name: str, threshold: float = 0.6) -> tuple:
        """
        Find best matching customer folder for a source name.
        Returns (matched_folder_name, similarity_score) or (None, 0) if no match.
        """
        if not source_name or not self.available_customers:
            return None, 0

        # Check manual aliases first
        aliases = self._load_customer_aliases()
        if source_name.upper() in aliases:
            alias_target = aliases[source_name.upper()]
            # Verify the alias target still exists
            if alias_target in self.available_customers:
                return alias_target, 1.0  # Perfect match via alias

        source_normalized = self._normalize_name(source_name)
        best_match = None
        best_score = 0

        for folder_name in self.available_customers:
            folder_normalized = self._normalize_name(folder_name)

            # Exact match after normalization
            if source_normalized == folder_normalized:
                return folder_name, 1.0

            # Check if one contains the other
            if source_normalized in folder_normalized or folder_normalized in source_normalized:
                score = 0.9
                if score > best_score:
                    best_score = score
                    best_match = folder_name
                continue

            # Fuzzy match using SequenceMatcher
            score = SequenceMatcher(None, source_normalized, folder_normalized).ratio()
            if score > best_score:
                best_score = score
                best_match = folder_name

        if best_score >= threshold:
            return best_match, best_score
        return None, 0

    def _detect_customer_column(self) -> str:
        """Detect which column contains customer names"""
        if self.source_df is None:
            return None

        # Common customer column names (case-insensitive check)
        customer_col_names = [
            'customer', 'customer name', 'customername', 'client',
            'client name', 'company', 'company name', 'account',
            'account name', 'sold to', 'sold-to', 'bill to', 'ship to',
            'customer id', 'customerid', 'cust', 'cust id', 'custid'
        ]

        for col in self.source_df.columns:
            col_lower = str(col).lower().strip()
            if col_lower in customer_col_names:
                return col

        return None

    # ==================== Report Processing ====================

    def fix_and_export(self):
        """Fix the report and export to customer's reports folder"""
        if self.source_df is None or self.template_columns is None:
            self.show_error("Missing Data", "Please load both template and source files first")
            return

        selected = self.customer_combo.currentText().strip()

        # Get customer files directory
        cf_dir = self.app_context.get_setting('customer_files_dir', '')
        if not cf_dir or not Path(cf_dir).exists():
            self.show_error("Directory Error", "Customer files directory not configured or doesn't exist")
            return

        self._log("=" * 50)

        # Check if auto-detect mode
        if selected == "-- Auto-detect from data --":
            self._export_multi_customer(cf_dir)
        else:
            self._export_single_customer(cf_dir, selected)

    def _find_last_report(self, reports_dir: Path, customer: str) -> Path:
        """Find the report file for a customer (single file, no date suffix)"""
        # Check for the standard report file (no date suffix)
        report_file = reports_dir / f"{customer}_jobRpt.xlsx"
        if report_file.exists():
            return report_file

        # Also check for legacy dated files and use most recent
        pattern = f"{customer}_jobRpt_*.xlsx"
        report_files = list(reports_dir.glob(pattern))
        if report_files:
            report_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            return report_files[0]

        return None

    def _get_completed_jobs(self, df_fixed: 'pd.DataFrame', reports_dir: Path, customer: str) -> tuple:
        """Compare with last report to:
        1. Preserve notes from previous report for jobs still active
        2. Mark missing jobs as complete
        3. Carry forward previously completed jobs
        4. Preserve highlighting from previous report

        Returns: (df_fixed, highlighted_rows) - highlighted_rows contains row indices that should be yellow
        """
        last_report = self._find_last_report(reports_dir, customer)
        if not last_report:
            self._log("No previous report found - skipping completion check")
            return df_fixed, set()

        try:
            df_previous = pd.read_excel(last_report)
            self._log(f"Comparing with previous report: {last_report.name}")
        except Exception as e:
            self._log(f"Could not read previous report: {e}")
            return df_fixed, set()

        # Read highlighting from previous report
        previous_highlighted_keys = set()
        wb_prev = None
        try:
            wb_prev = load_workbook(last_report)
            ws_prev = wb_prev.active

            # Find column indices first (outside the row loop)
            sched_col_prev = None
            po_col_prev = None
            line_col_prev = None
            for col_idx, cell in enumerate(ws_prev[1], 1):
                if cell.value == 'Scheduled End Date':
                    sched_col_prev = col_idx
                elif cell.value == 'Customer PO Number':
                    po_col_prev = col_idx
                elif cell.value == 'Line':
                    line_col_prev = col_idx

            if sched_col_prev and po_col_prev and line_col_prev:
                # Check each data row for yellow highlighting
                for row_idx in range(2, ws_prev.max_row + 1):
                    cell = ws_prev.cell(row=row_idx, column=sched_col_prev)

                    # Check for yellow fill - handle multiple ways openpyxl stores colors
                    is_yellow = False
                    if cell.fill and cell.fill.fill_type == 'solid':
                        fg_color = cell.fill.fgColor
                        if fg_color:
                            # Check RGB value (could be stored as rgb attribute or as index)
                            if fg_color.rgb:
                                rgb = str(fg_color.rgb).upper()
                                # Yellow is FFFF00, sometimes stored as 00FFFF00 (with alpha)
                                if 'FFFF00' in rgb:
                                    is_yellow = True
                            elif fg_color.theme is None and fg_color.indexed:
                                # Indexed color 13 is often yellow
                                if fg_color.indexed == 13:
                                    is_yellow = True

                    if is_yellow:
                        po_val = ws_prev.cell(row=row_idx, column=po_col_prev).value
                        line_val = ws_prev.cell(row=row_idx, column=line_col_prev).value
                        if po_val is not None and line_val is not None:
                            # Normalize values
                            po_str = str(po_val).strip()
                            line_str = str(line_val).strip()
                            if po_str.endswith('.0'):
                                po_str = po_str[:-2]
                            if line_str.endswith('.0'):
                                line_str = line_str[:-2]
                            previous_highlighted_keys.add(f"{po_str}|{line_str}")

            if previous_highlighted_keys:
                self._log(f"Found {len(previous_highlighted_keys)} highlighted cells in previous report")
            else:
                self._log("No highlighted cells found in previous report")
        except Exception as e:
            self._log(f"Could not read highlighting from previous report: {e}")
        finally:
            if wb_prev is not None:
                wb_prev.close()

        # Determine key columns for job identification
        use_composite_key = False
        if 'Job ID' in df_fixed.columns and 'Job ID' in df_previous.columns:
            key_cols = ['Job ID']
        elif 'Customer PO Number' in df_fixed.columns and 'Customer PO Number' in df_previous.columns:
            if 'Line' in df_fixed.columns and 'Line' in df_previous.columns:
                key_cols = ['Customer PO Number', 'Line']
                use_composite_key = True
            else:
                key_cols = ['Customer PO Number']
        else:
            self._log("Cannot identify job key column - skipping completion check")
            return df_fixed, set()

        # Create job key function - normalize values to handle int/float/string differences
        def normalize_value(val):
            """Normalize a value to string, handling int/float conversion"""
            if pd.isna(val):
                return ''
            # Convert to string
            s = str(val).strip()
            # Remove .0 suffix from floats that are actually integers
            if s.endswith('.0'):
                s = s[:-2]
            return s

        def make_key(row):
            if use_composite_key:
                po = normalize_value(row['Customer PO Number'])
                line = normalize_value(row['Line'])
                return f"{po}|{line}"
            else:
                return normalize_value(row[key_cols[0]])

        # Build map of current jobs (key -> row index in df_fixed)
        current_job_map = {}
        for idx, row in df_fixed.iterrows():
            try:
                key = make_key(row)
                if key and key not in ('nan|nan', 'nan', ''):
                    current_job_map[key] = idx
            except Exception as e:
                self._log(f"Warning: could not build key for current row {idx}: {e}")

        current_jobs = set(current_job_map.keys())
        self._log(f"Current report has {len(current_jobs)} unique jobs")

        # Build map of previous jobs (key -> row data)
        previous_job_map = {}  # key -> row index
        previous_notes = {}  # key -> notes value
        previous_completed = set()  # keys that were marked Complete
        previous_active = set()  # keys that were NOT marked Complete

        for idx, row in df_previous.iterrows():
            try:
                key = make_key(row)
                if key and key not in ('nan|nan', 'nan', ''):
                    previous_job_map[key] = idx
                    # Store notes from previous report
                    if 'Notes' in df_previous.columns:
                        notes_val = row.get('Notes', '')
                        if pd.notna(notes_val) and str(notes_val).strip():
                            previous_notes[key] = str(notes_val).strip()
                    status = str(row.get('Status', '')).lower()
                    if 'complete' in status:
                        previous_completed.add(key)
                    else:
                        previous_active.add(key)
            except Exception as e:
                self._log(f"Warning: could not build key for previous row {idx}: {e}")

        self._log(f"Previous report: {len(previous_active)} active, {len(previous_completed)} completed")

        # STEP 1: Merge notes from previous report into current data for matching jobs
        if 'Notes' in df_fixed.columns and previous_notes:
            notes_merged = 0
            for key, prev_notes in previous_notes.items():
                if key in current_job_map:
                    idx = current_job_map[key]
                    current_notes = df_fixed.at[idx, 'Notes']
                    # Only copy if current notes are empty and previous has content
                    if pd.isna(current_notes) or str(current_notes).strip() == '':
                        df_fixed.at[idx, 'Notes'] = prev_notes
                        notes_merged += 1
            if notes_merged > 0:
                self._log(f"Merged {notes_merged} notes from previous report")

        # STEP 2: Find jobs to add (completed)
        # Previously active jobs that are now missing = newly completed
        newly_completed = previous_active - current_jobs

        # Previously completed jobs that are still missing = carry forward
        carry_forward = previous_completed - current_jobs

        self._log(f"Newly completed: {len(newly_completed)}, Carry forward: {len(carry_forward)}")

        if not newly_completed and not carry_forward:
            self._log("No completed jobs to add")
            return df_fixed, previous_highlighted_keys

        rows_to_add = []
        original_columns = list(df_fixed.columns)

        # Process newly completed jobs
        if newly_completed:
            self._log(f"Marking {len(newly_completed)} jobs as newly completed")
            indices = [previous_job_map[k] for k in newly_completed if k in previous_job_map]
            newly_completed_rows = df_previous.loc[indices].copy()

            # Mark as Complete with date
            newly_completed_rows['Status'] = 'Complete'
            if 'Notes' in newly_completed_rows.columns:
                date_str = datetime.now().strftime('%m/%d')
                newly_completed_rows['Notes'] = newly_completed_rows['Notes'].fillna('').astype(str).apply(
                    lambda x: f"{x} [Completed {date_str}]".strip()
                )

            rows_to_add.append(newly_completed_rows)

        # Carry forward previously completed jobs
        if carry_forward:
            self._log(f"Carrying forward {len(carry_forward)} previously completed jobs")
            indices = [previous_job_map[k] for k in carry_forward if k in previous_job_map]
            carry_rows = df_previous.loc[indices].copy()
            rows_to_add.append(carry_rows)

        # Combine and clean up
        if rows_to_add:
            completed_rows = pd.concat(rows_to_add, ignore_index=True)

            # Remove any leftover _job_key column
            if '_job_key' in completed_rows.columns:
                completed_rows.drop('_job_key', axis=1, inplace=True)

            # Match columns to df_fixed
            for col in original_columns:
                if col not in completed_rows.columns:
                    completed_rows[col] = None

            # Only keep expected columns
            completed_rows = completed_rows[[col for col in original_columns if col in completed_rows.columns]]

            # Append to current data
            df_combined = pd.concat([df_fixed, completed_rows], ignore_index=True)

            self._log(f"Added {len(completed_rows)} completed jobs to report")
            return df_combined, previous_highlighted_keys

        return df_fixed, previous_highlighted_keys

    def _export_single_customer(self, cf_dir: str, customer: str):
        """Export report for a single selected customer"""
        customer_path = Path(cf_dir) / customer
        if not customer_path.exists():
            self.show_error("Customer Error", f"Customer folder not found:\n{customer_path}")
            return

        # Create reports directory if needed
        reports_dir = customer_path / "reports"
        reports_dir.mkdir(exist_ok=True)

        # Build output filename (overwrites previous report)
        output_file = reports_dir / f"{customer}_jobRpt.xlsx"

        self._log("Starting single-customer report transformation...")

        try:
            # Apply transformation
            df_fixed = self._transform_report(self.source_df)

            # Filter out sub-jobs (Job IDs ending in a letter)
            df_fixed = self._filter_letter_suffix_jobs(df_fixed)

            # Track schedule changes after filtering so row indices are correct
            changed_rows = self._track_schedule_changes(df_fixed)

            # Check for completed jobs from previous report (also returns preserved highlights)
            df_fixed, prev_highlighted_keys = self._get_completed_jobs(df_fixed, reports_dir, customer)

            # Save with formatting (merge new changes with preserved highlights)
            self._save_formatted_excel(df_fixed, output_file, changed_rows, prev_highlighted_keys)

            self.last_output_path = output_file
            self.open_output_btn.setEnabled(True)

            self._log("=" * 50)
            self._log(f"SUCCESS: Report saved to:")
            self._log(str(output_file))
            self._log("=" * 50)

            self.show_info("Export Complete",
                          f"Report saved to:\n{output_file}\n\n"
                          f"Schedule changes: {len(changed_rows)}")

        except Exception as e:
            self._log(f"ERROR: {str(e)}")
            self.show_error("Export Failed", f"Failed to export report:\n{str(e)}")

    def _export_multi_customer(self, cf_dir: str):
        """Export separate reports for each customer in the data"""
        self._log("Starting multi-customer report transformation...")

        # Detect customer column
        customer_col = self._detect_customer_column()
        if not customer_col:
            self.show_error("No Customer Column",
                          "Could not detect a customer column in the source data.\n\n"
                          "Expected column names: Customer, Customer Name, Client, Company, etc.\n\n"
                          "Please select a specific customer from the dropdown instead.")
            return

        self._log(f"Detected customer column: '{customer_col}'")

        # Use cached mapping if available from preview, otherwise compute it
        if self.customer_mapping:
            self._log("Using cached customer mapping from preview")
            # Convert from (folder, score, rows) to (folder, score) format
            customer_mapping = {k: (v[0], v[1]) for k, v in self.customer_mapping.items()}
            unmatched = [name for name, _ in self.unmatched_customers]
        else:
            # Get unique customers from data
            unique_customers = self.source_df[customer_col].dropna().unique()
            self._log(f"Found {len(unique_customers)} unique customers in data")

            # Match each customer to a folder
            customer_mapping = {}
            unmatched = []

            for src_customer in unique_customers:
                src_name = str(src_customer).strip()
                if not src_name:
                    continue

                matched_folder, score = self._fuzzy_match_customer(src_name)
                if matched_folder:
                    customer_mapping[src_name] = (matched_folder, score)
                    self._log(f"  '{src_name}' -> '{matched_folder}' (score: {score:.2f})")
                else:
                    unmatched.append(src_name)
                    self._log(f"  '{src_name}' -> NO MATCH FOUND")

        # Warn about unmatched customers
        if unmatched:
            msg = f"Could not match {len(unmatched)} customer(s) to folders:\n\n"
            msg += "\n".join(f"  - {name}" for name in unmatched[:10])
            if len(unmatched) > 10:
                msg += f"\n  ... and {len(unmatched) - 10} more"
            msg += "\n\nThese customers will be skipped. Continue?"

            reply = QMessageBox.question(
                self._widget, "Unmatched Customers", msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                self._log("Export cancelled by user")
                return

        if not customer_mapping:
            self.show_error("No Matches", "Could not match any customers to folders.")
            return

        # Process each customer
        success_count = 0
        total_changes = 0
        exported_files = []

        for src_name, (folder_name, _) in customer_mapping.items():
            self._log("-" * 40)
            self._log(f"Processing: {src_name} -> {folder_name}")

            # Filter data for this customer
            customer_df = self.source_df[
                self.source_df[customer_col].astype(str).str.strip() == src_name
            ].copy()

            if customer_df.empty:
                self._log(f"  No rows found, skipping")
                continue

            self._log(f"  {len(customer_df)} rows")

            # Setup output path (overwrites previous report)
            customer_path = Path(cf_dir) / folder_name
            reports_dir = customer_path / "reports"
            reports_dir.mkdir(exist_ok=True)
            output_file = reports_dir / f"{folder_name}_jobRpt.xlsx"

            try:
                # Transform this customer's data
                df_fixed = self._transform_report(customer_df)

                # Filter out sub-jobs (Job IDs ending in a letter)
                df_fixed = self._filter_letter_suffix_jobs(df_fixed)

                # Track schedule changes after filtering so row indices are correct
                changed_rows = self._track_schedule_changes(df_fixed)

                # Check for completed jobs from previous report (also returns preserved highlights)
                df_fixed, prev_highlighted_keys = self._get_completed_jobs(df_fixed, reports_dir, folder_name)

                # Save with formatting (merge new changes with preserved highlights)
                self._save_formatted_excel(df_fixed, output_file, changed_rows, prev_highlighted_keys)

                success_count += 1
                total_changes += len(changed_rows)
                exported_files.append(str(output_file))
                self._log(f"  Saved: {output_file.name} ({len(changed_rows)} changes)")

            except Exception as e:
                self._log(f"  ERROR: {str(e)}")

        # Final summary
        self._log("=" * 50)
        self._log(f"COMPLETE: {success_count}/{len(customer_mapping)} reports exported")
        self._log(f"Total schedule changes: {total_changes}")
        if unmatched:
            self._log(f"Skipped (no match): {len(unmatched)}")
        self._log("=" * 50)

        # Enable open folder button for first exported file
        if exported_files:
            self.last_output_path = Path(exported_files[0])
            self.open_output_btn.setEnabled(True)

        # Show summary
        summary = f"Exported {success_count} report(s)\n"
        summary += f"Total schedule changes: {total_changes}\n"
        if unmatched:
            summary += f"Skipped (unmatched): {len(unmatched)}\n"
        summary += f"\nFiles saved to customer reports folders."

        self.show_info("Multi-Customer Export Complete", summary)

    def _extract_dpas_ratings(self, source_df: 'pd.DataFrame') -> list:
        """Scan all source columns for DPAS rating patterns (e.g. DX-A3, DO-B1).
        Returns a positional list aligned with source_df rows."""
        pattern = re.compile(r'\bD[OX]-[A-Z]\d+\b', re.IGNORECASE)
        ratings = [''] * len(source_df)
        for col in source_df.columns:
            col_values = source_df[col].astype(str).tolist()
            for i, val in enumerate(col_values):
                if not ratings[i]:
                    match = pattern.search(val)
                    if match:
                        ratings[i] = match.group(0).upper()
        return ratings

    def _filter_letter_suffix_jobs(self, df: 'pd.DataFrame') -> 'pd.DataFrame':
        """Remove rows where Job ID ends in a letter (e.g. 12345A, 67890B).
        These are sub-jobs that should not appear in customer reports."""
        if 'Job ID' not in df.columns:
            return df

        before_count = len(df)
        # Keep rows where Job ID is NaN, empty, or does NOT end with a letter
        mask = df['Job ID'].apply(
            lambda x: pd.isna(x) or not re.search(r'[A-Za-z]$', str(x).strip())
        )
        df_filtered = df[mask].reset_index(drop=True)
        removed = before_count - len(df_filtered)
        if removed > 0:
            self._log(f"Filtered out {removed} sub-jobs (Job ID ending in a letter)")
        return df_filtered

    def _transform_report(self, source_df: 'pd.DataFrame'):
        """Transform source data to match template layout"""
        df_fixed = pd.DataFrame()

        # Copy/create columns in template order
        for col in self.template_columns:
            if col in source_df.columns:
                df_fixed[col] = source_df[col].values
            else:
                df_fixed[col] = None
                self._log(f"Added empty column: {col}")

        removed_count = len(set(source_df.columns) - set(self.template_columns))
        self._log(f"Removed {removed_count} columns not in template")
        self._log(f"Result: {len(df_fixed)} rows x {len(df_fixed.columns)} columns")

        # Merge Promise Date from delivery schedule (if loaded)
        if self.delivery_df is not None and 'Job ID' in df_fixed.columns:
            df_fixed['_job_id_str'] = df_fixed['Job ID'].astype(str).str.strip()
            # Drop the empty placeholder added from template to avoid duplicate column conflict
            if 'Promise Date' in df_fixed.columns:
                df_fixed = df_fixed.drop(columns=['Promise Date'])
            df_fixed = df_fixed.merge(
                self.delivery_df,
                left_on='_job_id_str',
                right_on='_delivery_job_id',
                how='left'
            )
            df_fixed.drop(columns=['_job_id_str', '_delivery_job_id'], inplace=True)
            matched = df_fixed['Promise Date'].notna().sum()
            self._log(f"Promise Date merged: {matched}/{len(df_fixed)} rows matched")

            # Reposition Promise Date immediately after Scheduled End Date
            if 'Scheduled End Date' in df_fixed.columns and 'Promise Date' in df_fixed.columns:
                cols = list(df_fixed.columns)
                cols.remove('Promise Date')
                insert_at = cols.index('Scheduled End Date') + 1
                cols.insert(insert_at, 'Promise Date')
                df_fixed = df_fixed[cols]

        # Process Scheduled End Date per PO
        if 'Customer PO Number' in df_fixed.columns and 'Scheduled End Date' in df_fixed.columns:
            self._log("Processing Scheduled End Date per PO...")

            # Convert to datetime first (keep as datetime for groupby max to work)
            df_fixed['Scheduled End Date'] = pd.to_datetime(
                df_fixed['Scheduled End Date'], errors='coerce'
            )

            # Set all rows to max date per PO (works with datetime64)
            po_max_dates = df_fixed.groupby('Customer PO Number')['Scheduled End Date'].transform('max')
            df_fixed['Scheduled End Date'] = po_max_dates

            # Now convert to date only (after aggregation)
            df_fixed['Scheduled End Date'] = df_fixed['Scheduled End Date'].dt.date

            unique_pos = df_fixed['Customer PO Number'].nunique()
            self._log(f"Updated dates for {unique_pos} unique POs")

        # Extract DPAS ratings and populate Classification column
        dpas_ratings = self._extract_dpas_ratings(source_df)
        dpas_found = sum(1 for r in dpas_ratings if r)
        if dpas_found > 0:
            self._log(f"Detected {dpas_found} DPAS rating(s) — populating Classification column")
            if 'Classification' not in df_fixed.columns:
                df_fixed['Classification'] = ''
            col_vals = df_fixed['Classification'].tolist()
            for i, rating in enumerate(dpas_ratings):
                if rating:
                    existing = col_vals[i]
                    if existing is None or str(existing).strip() in ('', 'nan', 'None'):
                        col_vals[i] = rating
            df_fixed['Classification'] = col_vals

        return df_fixed

    def _track_schedule_changes(self, df_fixed):
        """Track schedule changes and add notes"""
        history_file = get_config_dir() / 'schedule_history.json'
        history = {}

        if history_file.exists():
            try:
                with open(history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
                # Prune entries not updated in the last 180 days
                cutoff = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d %H:%M:%S')
                history = {
                    k: v for k, v in history.items()
                    if v.get('last_updated', '') >= cutoff
                }
            except Exception:
                history = {}

        # Ensure Notes column exists
        if 'Notes' not in df_fixed.columns:
            df_fixed['Notes'] = ''

        changed_rows = []
        changes_found = 0

        for idx, row in df_fixed.iterrows():
            po_raw = row.get('Customer PO Number', '')
            line = str(row.get('Line', ''))
            current_date = row.get('Scheduled End Date')

            if pd.isna(po_raw) or str(po_raw).strip() == '' or pd.isna(current_date):
                continue
            po = str(po_raw)

            key = f"{po}|{line}"
            current_date_str = str(current_date) if current_date else ''

            if key in history:
                previous_date_str = history[key].get('scheduled_end_date', '')

                if previous_date_str and previous_date_str != current_date_str:
                    changes_found += 1
                    changed_rows.append(idx)

                    existing_notes = str(row.get('Notes', '')) if pd.notna(row.get('Notes')) else ''
                    change_note = f"[{datetime.now().strftime('%m/%d')}] Moved from {previous_date_str}"

                    if existing_notes:
                        new_notes = f"{change_note}; {existing_notes}"
                    else:
                        new_notes = change_note

                    df_fixed.at[idx, 'Notes'] = new_notes

            # Update history
            history[key] = {
                'scheduled_end_date': current_date_str,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'po': po,
                'line': line
            }

        # Save updated history
        try:
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, indent=2)
            self._log(f"History saved ({len(history)} entries)")
        except Exception as e:
            self._log(f"Warning: Could not save history: {e}")

        if changes_found > 0:
            self._log(f"Found {changes_found} schedule changes (will highlight yellow)")

        return changed_rows

    def _save_formatted_excel(self, df_fixed, output_file, changed_rows, prev_highlighted_keys=None):
        """Save DataFrame with Excel formatting

        Args:
            df_fixed: DataFrame to save
            output_file: Path to save to
            changed_rows: List of row indices with NEW schedule changes (from this run)
            prev_highlighted_keys: Set of job keys (PO|Line) that were highlighted in previous report
        """
        # Save initial Excel
        df_fixed.to_excel(output_file, index=False, engine='openpyxl')

        # Load and format
        wb = load_workbook(output_file)
        ws = wb.active
        if ws is None:
            self._log("Warning: workbook has no active sheet, skipping formatting")
            return

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Find column indices we need
        sched_col = None
        promise_col = None
        po_col = None
        line_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Scheduled End Date':
                sched_col = col_idx
            elif cell.value == 'Promise Date':
                promise_col = col_idx
            elif cell.value == 'Customer PO Number':
                po_col = col_idx
            elif cell.value == 'Line':
                line_col = col_idx

        highlighted_count = 0

        # Apply highlighting for NEW changed rows (from this run)
        if changed_rows and sched_col:
            for row_idx in changed_rows:
                excel_row = row_idx + 2  # 1-indexed + header
                ws.cell(row=excel_row, column=sched_col).fill = yellow_fill
                highlighted_count += 1

        # Apply highlighting for PRESERVED rows (from previous report)
        if prev_highlighted_keys and sched_col and po_col and line_col:
            # Normalize function for consistent key matching
            def normalize_value(val):
                if val is None:
                    return ''
                s = str(val).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                return s

            # Check each data row to see if it should be highlighted
            for excel_row in range(2, ws.max_row + 1):
                po_val = ws.cell(row=excel_row, column=po_col).value
                line_val = ws.cell(row=excel_row, column=line_col).value

                if po_val is not None and line_val is not None:
                    key = f"{normalize_value(po_val)}|{normalize_value(line_val)}"
                    if key in prev_highlighted_keys:
                        # Check if this row isn't already highlighted from new changes
                        # (changed_rows are 0-indexed, excel_row is 1-indexed+header)
                        row_idx = excel_row - 2
                        if row_idx not in changed_rows:
                            ws.cell(row=excel_row, column=sched_col).fill = yellow_fill
                            highlighted_count += 1

        if highlighted_count > 0:
            self._log(f"Highlighted {highlighted_count} cells (new changes + preserved)")

        # Highlight red where Scheduled End Date is after Promise Date (overrides yellow)
        late_count = 0
        if sched_col and promise_col:
            for excel_row in range(2, ws.max_row + 1):
                sched_val = ws.cell(row=excel_row, column=sched_col).value
                promise_val = ws.cell(row=excel_row, column=promise_col).value
                if sched_val and promise_val:
                    try:
                        sched_date = pd.to_datetime(sched_val).date()
                        promise_date = pd.to_datetime(promise_val).date()
                        if sched_date > promise_date:
                            ws.cell(row=excel_row, column=sched_col).fill = red_fill
                            ws.cell(row=excel_row, column=promise_col).fill = red_fill
                            late_count += 1
                    except Exception:
                        pass
            if late_count > 0:
                self._log(f"Highlighted {late_count} rows red (Scheduled End Date after Promise Date)")

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add table formatting
        max_row = ws.max_row
        max_col = ws.max_column
        table_ref = f"A1:{ws.cell(max_row, max_col).coordinate}"

        # Use sheet title in table name to avoid collisions if the file is re-opened
        safe_title = re.sub(r'[^A-Za-z0-9_]', '_', str(ws.title) if ws.title else 'Sheet')
        table = Table(displayName=f"Table_{safe_title}", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        wb.save(output_file)
        self._log("Applied formatting: auto-fit, frozen header, table style")

    def open_output_folder(self):
        """Open the output folder in file explorer"""
        if self.last_output_path and self.last_output_path.parent.exists():
            success, error = open_folder(str(self.last_output_path.parent))
            if not success:
                self.show_error("Error", error or "Could not open folder")

    def _log(self, message: str):
        """Add message to status log"""
        if self.status_text:
            self.status_text.append(message)
            # Scroll to bottom
            scrollbar = self.status_text.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())
        self.log_message(message)

    def cleanup(self):
        """Cleanup resources"""
        pass
